# %%
import json
from datetime import datetime
from numbers import Number

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


# sys.path.append(os.path.join("..", "PorttalUsuario", "remover_acesso", "Relação_contratos.xlsx"))
def build_date(date):
    if isinstance(date, Number):
        return datetime.fromtimestamp(date / 1000).strftime('%Y-%m-%dT%H:%M:%SZ')
    else:
        return datetime.fromisoformat(date).strftime('%Y-%m-%dT%H:%M:%SZ')


def csv_to_json(filename):
    wb = load_workbook(filename=filename)
    ws = wb.active

    my_list = []

    last_column = len(list(ws.columns))
    last_row = len(list(ws.rows))

    for row in range(1, last_row + 1):
        my_dict = {}
        for column in range(1, last_column + 1):
            column_letter = get_column_letter(column)
            if row > 1:
                column_letter_plus_row = ws[column_letter + str(1)].value
                column_title = column_letter + str(row)
                column_value = ws[column_title].value
                if isinstance(column_value, datetime):
                    # print("column_value", column_value)
                    column_value = build_date(str(column_value))
                # print("column_letter_plus_row",column_letter_plus_row)
                # print("column_title",column_title)
                # print("column_value",column_value)
                my_dict[column_letter_plus_row] = column_value
        my_list.append(my_dict)

    data = json.dumps(my_list, sort_keys=True, indent=4)  # with open('D:/data.json', 'w', encoding='utf-8') as f:
    return json.loads(data)


def do_it():
    data = csv_to_json('/Users/Mr-i-me/code/Mr-i-me-pontte/portal-scripts/layers/helpers/CARTORIOS DA BASE PONTTE.xlsx')

    with open('/Users/Mr-i-me/code/Mr-i-me-pontte/portal-scripts/layers/helpers/data.json', 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=4)
